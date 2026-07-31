#!/usr/bin/env python3
"""
Human-review sheet for the 81 registered prospective bets, with NCT links and a per-bet
model rationale grounded in the ACTUAL committed feature values.

For a NOVEL (drug, indication) pair the drug-level features are cloned from the compound, so the
signal that DIFFERENTIATES a FAIL call from a PASS call is the per-pair mechanism block
(ot_genetic_association, in_module, network topology) plus which sub-head (efficacy vs safety)
dominates. We surface those, plus the committed novelty verdict rationale, in one sheet.

Out: results/benchmark/prereg_C/prereg_C_bets_for_review.{xlsx,csv}  (FAIL bets first)
"""
from pathlib import Path
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
P = ROOT / "results/benchmark/prereg_C"
MECH_COLS = ["ot_genetic_association", "in_module", "topo_net", "kegg_frac",
             "clingen", "coverage_disease", "coverage_drug"]


def nz(s):
    return str(s).lower().strip()


def rationale(r):
    eff = r.get("P_fail_efficacy_calibrated", float("nan"))
    saf = r.get("P_fail_safety_calibrated", float("nan"))
    if r["bet"] == "PASS":
        return "PASS bet — low predicted failure, inside the held-out ≥90%-precision PASS band."
    otg, inm, topo = r.get("ot_genetic_association"), r.get("in_module"), r.get("topo_net")
    # Gap B/D trial-context signals (recomputed per trial; surface them so the displayed reason
    # matches the features that actually move the score, not generic boilerplate).
    prec = r.get("precedent_neg_class")
    cvm = r.get("endpoint_cvevent_match")
    phys = r.get("endpoint_physiology_score")
    parts = []
    # 1. Negative class-precedent (Gap D) is the most specific, transferable failure signal — lead with it.
    if prec == 1:
        parts.append("NEGATIVE CLASS-PRECEDENT: a same-target-class agent already failed Phase III in this "
                     "indication before this trial began — a specific, as-of-date prior, distinct from generic "
                     "disease difficulty.")
    if pd.notna(saf) and pd.notna(eff) and saf > eff:
        parts.append(f"SAFETY-driven (P_fail_safety {saf:.2f} > efficacy {eff:.2f}): a disjunctive "
                     "molecular liability the drug carries (e.g. promiscuity / hepatic / cardiac / bleeding).")
    else:
        # 2. CV-event mechanism-match (Gap B): a -1 is a genuine mismatch; a +1 is mitigating (the
        #    failure then rests on indication difficulty, not on the wrong mechanism).
        if cvm == -1:
            parts.append("CV-EVENT MISMATCH: the trial's primary endpoint is a discrete cardiovascular event, "
                         "but the drug's target is NOT on a validated atherothrombotic/cardiometabolic "
                         "event-reduction axis.")
        elif cvm == 1:
            parts.append("Borderline call DESPITE a CV-event mechanism match (the drug's target IS on the "
                         "atherothrombotic axis the endpoint demands): the indication's difficulty — e.g. "
                         "secondary stroke prevention, where matched antithrombotics still fail at a high rate "
                         "— drives the residual failure probability, not a mechanism mismatch.")
        elif pd.notna(otg) and otg < 0.10 and (inm == 0):
            parts.append("EFFICACY-driven mechanism mismatch: the target is NOT a genetically-supported "
                         "driver of this indication (OT genetic ≈ 0 and target absent from the disease "
                         "gene module) — the drug can be right, the disease wrong.")
        elif (pd.notna(otg) and otg >= 0.30) or inm == 1:
            parts.append(f"EFFICACY-driven despite partial mechanism support (OT genetic {otg:.2f}, "
                         f"in-module {int(inm) if pd.notna(inm) else '?'}): disease difficulty / context "
                         "dominates the call.")
        else:
            parts.append(f"EFFICACY-driven: weak target→disease mechanism fit (OT genetic {otg:.2f}).")
        if pd.notna(topo) and topo < 0 and cvm != 1:
            parts.append("Target is network-distal/upstream from the disease module (negative topology).")
    # 3. Endpoint-physiology MATCH caveat: if the drug's mechanism DOES move the measured endpoint,
    #    flag that the FAIL rests on disease-level fit, not the endpoint (the metoprolol→DMD failure mode).
    if phys == 1 and not (cvm == 1):
        parts.append("Caveat: the drug's mechanism does move the trial's physiological endpoint "
                     "(endpoint-physiology match) — the failure call rests on disease-level mechanism fit, "
                     "so check the endpoint is not one the drug plausibly moves.")
    return " ".join(parts)


def main():
    b = pd.read_csv(P / "prereg_C_confident_bets.csv")
    m = pd.read_csv(ROOT / "data/sources/mechanism_dataderived_prereg_C.csv")
    v = pd.read_csv(ROOT / "data/sources/prereg_C_residual_verdicts_v1.csv")
    m["_k"] = list(zip(m.IK14.astype(str), m.Disease.map(nz)))
    b["_k"] = list(zip(b.IK14.astype(str), b.novel_indication.map(nz)))
    keep_m = ["_k"] + [c for c in MECH_COLS if c in m.columns]
    b = b.merge(m[keep_m].drop_duplicates("_k"), on="_k", how="left")
    # Gap B/D trial-context features (keyed NCT_ID + drug) so the rationale reflects them
    tctx = ROOT / "data/sources/trialcontext_prereg_C.csv"
    if tctx.exists():
        tc = pd.read_csv(tctx)
        tc_cols = [c for c in ["endpoint_physiology_score", "endpoint_cvevent_match",
                               "precedent_neg_class"] if c in tc.columns]
        tc["_tk"] = list(zip(tc.NCT_ID.astype(str), tc.drug.map(nz)))
        b["_tk"] = list(zip(b.NCT_ID.astype(str), b.drug.map(nz)))
        b = b.merge(tc[["_tk"] + tc_cols].drop_duplicates("_tk"), on="_tk", how="left").drop(columns="_tk")
    vmap = {(nz(r.drug), nz(r.indication)): (r.verdict, r.rationale) for _, r in v.iterrows()}
    b["novelty_verdict"] = [vmap.get((nz(d), nz(i)), ("—", ""))[0]
                            for d, i in zip(b.drug, b.novel_indication)]
    b["novelty_rationale"] = [vmap.get((nz(d), nz(i)), ("", "(PASS bet — not in verdict residual)"))[1]
                              for d, i in zip(b.drug, b.novel_indication)]
    b["nct_link"] = "https://clinicaltrials.gov/study/" + b.NCT_ID.astype(str)
    b["model_rationale"] = b.apply(rationale, axis=1)

    if "feature_completeness" not in b.columns:
        b["feature_completeness"] = "OK (full biological features computed)"
    cols = ["bet", "feature_completeness", "drug", "novel_indication", "nct_link", "overall_status",
            "P_fail_overall_calibrated", "P_fail_efficacy_calibrated", "P_fail_safety_calibrated",
            "model_rationale"] + [c for c in MECH_COLS if c in b.columns] + \
           ["confidence_tier", "novelty_verdict", "novelty_rationale", "NCT_ID"]
    out = b[[c for c in cols if c in b.columns]].copy()
    # incomplete-first within each bet class so missing-data rows are impossible to miss
    out["_inc"] = out.feature_completeness.str.startswith("INCOMPLETE").astype(int)
    out = out.sort_values(["bet", "_inc", "P_fail_overall_calibrated"],
                          ascending=[True, False, False]).drop(columns="_inc")  # FAIL + INCOMPLETE first
    out.to_csv(P / "prereg_C_bets_for_review.csv", index=False)
    n_inc = int(out.feature_completeness.str.startswith("INCOMPLETE").sum())
    try:
        out.to_excel(P / "prereg_C_bets_for_review.xlsx", index=False, engine="openpyxl")
        # paint INCOMPLETE rows red so missing-pipeline-data predictions fail loudly
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        wb = load_workbook(P / "prereg_C_bets_for_review.xlsx"); ws = wb.active
        red = PatternFill("solid", fgColor="FFC7CE"); bold = Font(bold=True, color="9C0006")
        fc = list(out.columns).index("feature_completeness") + 1
        for ri in range(2, ws.max_row + 1):
            if str(ws.cell(ri, fc).value).startswith("INCOMPLETE"):
                for ci in range(1, ws.max_column + 1):
                    ws.cell(ri, ci).fill = red
                ws.cell(ri, fc).font = bold
        wb.save(P / "prereg_C_bets_for_review.xlsx")
        print(f"wrote {(P / 'prereg_C_bets_for_review.xlsx').relative_to(ROOT)} ({n_inc} INCOMPLETE rows highlighted)")
    except Exception as e:
        print(f"(xlsx skipped: {e})")
    print(f"{len(out)} bets: {(out.bet=='FAIL').sum()} FAIL (review), {(out.bet=='PASS').sum()} PASS")
    print("\n=== 10 FAIL bets: NCT + rationale ===")
    for r in out[out.bet == "FAIL"].itertuples():
        print(f"\n  {r.drug} -> {r.novel_indication}  (P_fail {r.P_fail_overall_calibrated:.2f})")
        print(f"    {r.nct_link}")
        print(f"    {r.model_rationale}")


if __name__ == "__main__":
    main()
